from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi import status as http_status
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import openpyxl
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('SECRET_KEY', 'garment-manufacturing-secret-key-change-in-production')
ALGORITHM = "HS256"

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    role: str  # admin, production_manager, user
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str = "user"

class UserLogin(BaseModel):
    email: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

# Masters Models
class Buyer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: str
    email: str
    phone: str
    address: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BuyerCreate(BaseModel):
    name: str
    contact_person: str
    email: str
    phone: str
    address: str

class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: str
    email: str
    phone: str
    address: str
    material_type: str  # fabric, trims, accessories
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    name: str
    contact_person: str
    email: str
    phone: str
    address: str
    material_type: str

class RawMaterial(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    material_type: str  # fabric, trims, accessories
    unit: str  # meters, pieces, kg
    cost_per_unit: float
    supplier_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RawMaterialCreate(BaseModel):
    name: str
    code: str
    material_type: str
    unit: str
    cost_per_unit: float
    supplier_id: Optional[str] = None

class Color(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    hex_value: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ColorCreate(BaseModel):
    name: str
    code: str
    hex_value: Optional[str] = None

class Size(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    sort_order: int
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SizeCreate(BaseModel):
    name: str
    code: str
    sort_order: int

class Article(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    description: str
    buyer_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ArticleCreate(BaseModel):
    name: str
    code: str
    description: str
    buyer_id: Optional[str] = None

# Fabric Model
class Fabric(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_type: str  # DYED, GREIGE, ZIP
    count_const: str
    fabric_name: str
    composition: str
    add_description: str
    gsm: Optional[int] = None
    width: Optional[str] = None
    color: Optional[str] = None
    final_item: str
    avg_roll_size: Optional[str] = None
    unit: str
    image_url: Optional[str] = None  # New field for image
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FabricCreate(BaseModel):
    item_type: str
    count_const: str
    fabric_name: str
    composition: str
    add_description: str
    gsm: Optional[int] = None
    width: Optional[str] = None
    color: Optional[str] = None
    final_item: str
    avg_roll_size: Optional[str] = None
    unit: str
    image_url: Optional[str] = None  # New field for image

# BOM Models
class BOMItem(BaseModel):
    material_id: str
    material_name: str
    avg_consumption: float
    wastage_percent: float
    total_consumption: float
    cost_per_unit: float
    total_cost: float

class BOM(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    article_id: str
    article_name: str
    color_id: str
    color_name: str
    items: List[BOMItem]
    total_cost: float
    status: str = "unassigned"  # unassigned, assigned
    mrp_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BOMCreate(BaseModel):
    article_id: str
    color_id: str
    items: List[BOMItem]

# MRP Models
class MRPMaterialRequirement(BaseModel):
    material_id: str
    material_name: str
    material_code: str
    unit: str
    total_quantity: float
    cost_per_unit: float
    total_cost: float

class MRP(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    mrp_number: str
    bom_ids: List[str]
    material_requirements: List[MRPMaterialRequirement]
    total_cost: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class MRPCreate(BaseModel):
    bom_ids: List[str]

# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: timedelta = timedelta(days=7)):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + expires_delta
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
        user = await db.users.find_one({"email": email}, {"_id": 0, "password": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Auth Routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_input: UserCreate):
    # Check if user exists
    existing = await db.users.find_one({"email": user_input.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user_dict = user_input.model_dump()
    hashed_password = hash_password(user_dict.pop("password"))
    user_obj = User(**user_dict)
    
    doc = user_obj.model_dump()
    doc['password'] = hashed_password
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    
    # Create token
    access_token = create_access_token(data={"sub": user_obj.email})
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.post("/auth/login", response_model=Token)
async def login(user_input: UserLogin):
    user_doc = await db.users.find_one({"email": user_input.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(user_input.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user_doc.pop('password')
    if isinstance(user_doc['created_at'], str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user_obj = User(**user_doc)
    access_token = create_access_token(data={"sub": user_obj.email})
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Buyer Routes
@api_router.post("/buyers", response_model=Buyer)
async def create_buyer(buyer_input: BuyerCreate, current_user: User = Depends(get_current_user)):
    buyer_obj = Buyer(**buyer_input.model_dump())
    doc = buyer_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.buyers.insert_one(doc)
    return buyer_obj

@api_router.get("/buyers", response_model=List[Buyer])
async def get_buyers(current_user: User = Depends(get_current_user)):
    buyers = await db.buyers.find({}, {"_id": 0}).to_list(1000)
    for buyer in buyers:
        if isinstance(buyer['created_at'], str):
            buyer['created_at'] = datetime.fromisoformat(buyer['created_at'])
    return buyers

@api_router.put("/buyers/{buyer_id}", response_model=Buyer)
async def update_buyer(buyer_id: str, buyer_input: BuyerCreate, current_user: User = Depends(get_current_user)):
    result = await db.buyers.update_one({"id": buyer_id}, {"$set": buyer_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Buyer not found")
    updated = await db.buyers.find_one({"id": buyer_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Buyer(**updated)

@api_router.delete("/buyers/{buyer_id}")
async def delete_buyer(buyer_id: str, current_user: User = Depends(get_current_user)):
    result = await db.buyers.delete_one({"id": buyer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Buyer not found")
    return {"message": "Buyer deleted successfully"}

# Supplier Routes
@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_input: SupplierCreate, current_user: User = Depends(get_current_user)):
    supplier_obj = Supplier(**supplier_input.model_dump())
    doc = supplier_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.suppliers.insert_one(doc)
    return supplier_obj

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    for supplier in suppliers:
        if isinstance(supplier['created_at'], str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return suppliers

@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, supplier_input: SupplierCreate, current_user: User = Depends(get_current_user)):
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": supplier_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Supplier(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: User = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}

# Raw Material Routes
@api_router.post("/raw-materials", response_model=RawMaterial)
async def create_raw_material(material_input: RawMaterialCreate, current_user: User = Depends(get_current_user)):
    material_obj = RawMaterial(**material_input.model_dump())
    doc = material_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.raw_materials.insert_one(doc)
    return material_obj

@api_router.get("/raw-materials", response_model=List[RawMaterial])
async def get_raw_materials(current_user: User = Depends(get_current_user)):
    materials = await db.raw_materials.find({}, {"_id": 0}).to_list(1000)
    for material in materials:
        if isinstance(material['created_at'], str):
            material['created_at'] = datetime.fromisoformat(material['created_at'])
    return materials

@api_router.put("/raw-materials/{material_id}", response_model=RawMaterial)
async def update_raw_material(material_id: str, material_input: RawMaterialCreate, current_user: User = Depends(get_current_user)):
    result = await db.raw_materials.update_one({"id": material_id}, {"$set": material_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    updated = await db.raw_materials.find_one({"id": material_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return RawMaterial(**updated)

@api_router.delete("/raw-materials/{material_id}")
async def delete_raw_material(material_id: str, current_user: User = Depends(get_current_user)):
    result = await db.raw_materials.delete_one({"id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Raw material not found")
    return {"message": "Raw material deleted successfully"}

# Color Routes
@api_router.post("/colors", response_model=Color)
async def create_color(color_input: ColorCreate, current_user: User = Depends(get_current_user)):
    color_obj = Color(**color_input.model_dump())
    doc = color_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.colors.insert_one(doc)
    return color_obj

@api_router.get("/colors", response_model=List[Color])
async def get_colors(current_user: User = Depends(get_current_user)):
    colors = await db.colors.find({}, {"_id": 0}).to_list(1000)
    for color in colors:
        if isinstance(color['created_at'], str):
            color['created_at'] = datetime.fromisoformat(color['created_at'])
    return colors

@api_router.put("/colors/{color_id}", response_model=Color)
async def update_color(color_id: str, color_input: ColorCreate, current_user: User = Depends(get_current_user)):
    result = await db.colors.update_one({"id": color_id}, {"$set": color_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Color not found")
    updated = await db.colors.find_one({"id": color_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Color(**updated)

@api_router.delete("/colors/{color_id}")
async def delete_color(color_id: str, current_user: User = Depends(get_current_user)):
    result = await db.colors.delete_one({"id": color_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Color not found")
    return {"message": "Color deleted successfully"}

# Size Routes
@api_router.post("/sizes", response_model=Size)
async def create_size(size_input: SizeCreate, current_user: User = Depends(get_current_user)):
    size_obj = Size(**size_input.model_dump())
    doc = size_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.sizes.insert_one(doc)
    return size_obj

@api_router.get("/sizes", response_model=List[Size])
async def get_sizes(current_user: User = Depends(get_current_user)):
    sizes = await db.sizes.find({}, {"_id": 0}).sort("sort_order", 1).to_list(1000)
    for size in sizes:
        if isinstance(size['created_at'], str):
            size['created_at'] = datetime.fromisoformat(size['created_at'])
    return sizes

@api_router.put("/sizes/{size_id}", response_model=Size)
async def update_size(size_id: str, size_input: SizeCreate, current_user: User = Depends(get_current_user)):
    result = await db.sizes.update_one({"id": size_id}, {"$set": size_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Size not found")
    updated = await db.sizes.find_one({"id": size_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Size(**updated)

@api_router.delete("/sizes/{size_id}")
async def delete_size(size_id: str, current_user: User = Depends(get_current_user)):
    result = await db.sizes.delete_one({"id": size_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Size not found")
    return {"message": "Size deleted successfully"}

# Article Routes
@api_router.post("/articles", response_model=Article)
async def create_article(article_input: ArticleCreate, current_user: User = Depends(get_current_user)):
    article_obj = Article(**article_input.model_dump())
    doc = article_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.articles.insert_one(doc)
    return article_obj

@api_router.get("/articles", response_model=List[Article])
async def get_articles(current_user: User = Depends(get_current_user)):
    articles = await db.articles.find({}, {"_id": 0}).to_list(1000)
    for article in articles:
        if isinstance(article['created_at'], str):
            article['created_at'] = datetime.fromisoformat(article['created_at'])
    return articles

@api_router.put("/articles/{article_id}", response_model=Article)
async def update_article(article_id: str, article_input: ArticleCreate, current_user: User = Depends(get_current_user)):
    result = await db.articles.update_one({"id": article_id}, {"$set": article_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    updated = await db.articles.find_one({"id": article_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Article(**updated)

@api_router.delete("/articles/{article_id}")
async def delete_article(article_id: str, current_user: User = Depends(get_current_user)):
    result = await db.articles.delete_one({"id": article_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Article not found")
    return {"message": "Article deleted successfully"}

# Fabric Routes
@api_router.post("/fabrics", response_model=Fabric)
async def create_fabric(fabric_input: FabricCreate, current_user: User = Depends(get_current_user)):
    fabric_obj = Fabric(**fabric_input.model_dump())
    doc = fabric_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.fabrics.insert_one(doc)
    return fabric_obj

@api_router.get("/fabrics", response_model=List[Fabric])
async def get_fabrics(current_user: User = Depends(get_current_user)):
    fabrics = await db.fabrics.find({}, {"_id": 0}).to_list(1000)
    for fabric in fabrics:
        if isinstance(fabric['created_at'], str):
            fabric['created_at'] = datetime.fromisoformat(fabric['created_at'])
    return fabrics

@api_router.put("/fabrics/{fabric_id}", response_model=Fabric)
async def update_fabric(fabric_id: str, fabric_input: FabricCreate, current_user: User = Depends(get_current_user)):
    result = await db.fabrics.update_one({"id": fabric_id}, {"$set": fabric_input.model_dump()})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Fabric not found")
    updated = await db.fabrics.find_one({"id": fabric_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Fabric(**updated)

@api_router.delete("/fabrics/{fabric_id}")
async def delete_fabric(fabric_id: str, current_user: User = Depends(get_current_user)):
    result = await db.fabrics.delete_one({"id": fabric_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fabric not found")
    return {"message": "Fabric deleted successfully"}

# BOM Routes
@api_router.post("/boms/comprehensive")
async def create_comprehensive_bom(bom_data: dict, current_user: User = Depends(get_current_user)):
    try:
        # Extract header and all tabs data
        header = bom_data.get("header", {})
        fabric_tables = bom_data.get("fabricTables", [])
        trims_tables = bom_data.get("trimsTables", [])
        operations = bom_data.get("operations", [])
        
        # Create BOM document with all three tabs
        bom_id = str(uuid.uuid4())
        bom_doc = {
            "id": bom_id,
            "header": header,
            "fabricTables": fabric_tables,
            "trimsTables": trims_tables,
            "operations": operations,
            "status": "assigned",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.username
        }
        
        await db.comprehensive_boms.insert_one(bom_doc)
        
        return {
            "message": "BOM created successfully with all tabs",
            "bom_id": bom_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating BOM: {str(e)}")

@api_router.post("/boms", response_model=BOM)
async def create_bom(bom_input: BOMCreate, current_user: User = Depends(get_current_user)):
    # Get article and color names
    article = await db.articles.find_one({"id": bom_input.article_id}, {"_id": 0})
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    
    color = await db.colors.find_one({"id": bom_input.color_id}, {"_id": 0})
    if not color:
        raise HTTPException(status_code=404, detail="Color not found")
    
    # Calculate total cost
    total_cost = sum(item.total_cost for item in bom_input.items)
    
    bom_obj = BOM(
        article_id=bom_input.article_id,
        article_name=article['name'],
        color_id=bom_input.color_id,
        color_name=color['name'],
        items=bom_input.items,
        total_cost=total_cost
    )
    
    doc = bom_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.boms.insert_one(doc)
    return bom_obj

@api_router.get("/boms")
async def get_boms(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    
    # Fetch from both collections
    boms_regular = await db.boms.find(query, {"_id": 0}).to_list(1000)
    boms_comprehensive = await db.comprehensive_boms.find(query, {"_id": 0}).to_list(1000)
    
    # Process regular BOMs
    for bom in boms_regular:
        if isinstance(bom.get('created_at'), str):
            bom['created_at'] = datetime.fromisoformat(bom['created_at'])
        bom['bom_type'] = 'regular'
    
    # Process comprehensive BOMs
    for bom in boms_comprehensive:
        if isinstance(bom.get('created_at'), str):
            bom['created_at'] = datetime.fromisoformat(bom['created_at'])
        bom['bom_type'] = 'comprehensive'
    
    # Combine both lists
    all_boms = boms_regular + boms_comprehensive
    
    return all_boms

@api_router.get("/boms/{bom_id}")
async def get_bom(bom_id: str, current_user: User = Depends(get_current_user)):
    # Try finding in regular BOMs first
    bom = await db.boms.find_one({"id": bom_id}, {"_id": 0})
    
    # If not found, try comprehensive BOMs
    if not bom:
        bom = await db.comprehensive_boms.find_one({"id": bom_id}, {"_id": 0})
    
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")
    
    if isinstance(bom.get('created_at'), str):
        bom['created_at'] = datetime.fromisoformat(bom['created_at'])
    
    return bom

@api_router.put("/boms/{bom_id}")
async def update_bom(bom_id: str, bom_data: dict, current_user: User = Depends(get_current_user)):
    try:
        # Check if BOM exists in either collection
        existing_bom = await db.boms.find_one({"id": bom_id})
        collection = db.boms
        
        if not existing_bom:
            existing_bom = await db.comprehensive_boms.find_one({"id": bom_id})
            collection = db.comprehensive_boms
        
        if not existing_bom:
            raise HTTPException(status_code=404, detail="BOM not found")
        
        # Update the BOM
        header = bom_data.get("header", {})
        fabric_tables = bom_data.get("fabricTables", [])
        trims_tables = bom_data.get("trimsTables", [])
        operations = bom_data.get("operations", [])
        
        update_doc = {
            "header": header,
            "fabricTables": fabric_tables,
            "trimsTables": trims_tables,
            "operations": operations,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": current_user.username
        }
        
        result = await collection.update_one(
            {"id": bom_id},
            {"$set": update_doc}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=400, detail="BOM update failed")
        
        return {
            "message": "BOM updated successfully",
            "bom_id": bom_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating BOM: {str(e)}")

@api_router.delete("/boms/{bom_id}")
async def delete_bom(bom_id: str, current_user: User = Depends(get_current_user)):
    # Try deleting from both collections
    result1 = await db.boms.delete_one({"id": bom_id})
    result2 = await db.comprehensive_boms.delete_one({"id": bom_id})
    
    if result1.deleted_count == 0 and result2.deleted_count == 0:
        raise HTTPException(status_code=404, detail="BOM not found")
    return {"message": "BOM deleted successfully"}

# MRP Routes
@api_router.post("/mrps", response_model=MRP)
async def create_mrp(mrp_input: MRPCreate, current_user: User = Depends(get_current_user)):
    if not mrp_input.bom_ids:
        raise HTTPException(status_code=400, detail="At least one BOM must be selected")
    
    # Get all selected BOMs
    boms = await db.boms.find({"id": {"$in": mrp_input.bom_ids}, "status": "unassigned"}, {"_id": 0}).to_list(1000)
    
    if len(boms) != len(mrp_input.bom_ids):
        raise HTTPException(status_code=400, detail="Some BOMs not found or already assigned")
    
    # Generate MRP number
    mrp_count = await db.mrps.count_documents({})
    mrp_number = f"MRP-{mrp_count + 1:05d}"
    
    # Consolidate material requirements
    material_map = {}
    for bom in boms:
        for item in bom['items']:
            mat_id = item['material_id']
            if mat_id not in material_map:
                # Get material details
                material = await db.raw_materials.find_one({"id": mat_id}, {"_id": 0})
                material_map[mat_id] = {
                    "material_id": mat_id,
                    "material_name": item['material_name'],
                    "material_code": material['code'] if material else "",
                    "unit": material['unit'] if material else "",
                    "total_quantity": 0,
                    "cost_per_unit": item['cost_per_unit'],
                    "total_cost": 0
                }
            material_map[mat_id]["total_quantity"] += item['total_consumption']
            material_map[mat_id]["total_cost"] += item['total_cost']
    
    material_requirements = [MRPMaterialRequirement(**mat) for mat in material_map.values()]
    total_cost = sum(mat.total_cost for mat in material_requirements)
    
    # Create MRP
    mrp_obj = MRP(
        mrp_number=mrp_number,
        bom_ids=mrp_input.bom_ids,
        material_requirements=material_requirements,
        total_cost=total_cost,
        created_by=current_user.username
    )
    
    doc = mrp_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.mrps.insert_one(doc)
    
    # Update BOMs as assigned
    await db.boms.update_many(
        {"id": {"$in": mrp_input.bom_ids}},
        {"$set": {"status": "assigned", "mrp_id": mrp_obj.id}}
    )
    
    return mrp_obj

@api_router.get("/mrps", response_model=List[MRP])
async def get_mrps(current_user: User = Depends(get_current_user)):
    mrps = await db.mrps.find({}, {"_id": 0}).to_list(1000)
    for mrp in mrps:
        if isinstance(mrp['created_at'], str):
            mrp['created_at'] = datetime.fromisoformat(mrp['created_at'])
    return mrps

@api_router.get("/mrps/{mrp_id}", response_model=MRP)
async def get_mrp(mrp_id: str, current_user: User = Depends(get_current_user)):
    mrp = await db.mrps.find_one({"id": mrp_id}, {"_id": 0})
    if not mrp:
        raise HTTPException(status_code=404, detail="MRP not found")
    if isinstance(mrp['created_at'], str):
        mrp['created_at'] = datetime.fromisoformat(mrp['created_at'])
    return MRP(**mrp)

@api_router.delete("/mrps/{mrp_id}")
async def delete_mrp(mrp_id: str, current_user: User = Depends(get_current_user)):
    # Get MRP to unassign BOMs
    mrp = await db.mrps.find_one({"id": mrp_id}, {"_id": 0})
    if not mrp:
        raise HTTPException(status_code=404, detail="MRP not found")
    
    # Unassign BOMs
    await db.boms.update_many(
        {"id": {"$in": mrp['bom_ids']}},
        {"$set": {"status": "unassigned"}, "$unset": {"mrp_id": ""}}
    )
    
    # Delete MRP
    await db.mrps.delete_one({"id": mrp_id})
    return {"message": "MRP deleted successfully"}


# ============================================================================
# DYNAMIC MASTER BUILDER SYSTEM
# ============================================================================

# Models for Dynamic Master Builder
class FieldConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Field name
    label: str  # Display label
    type: str  # text, number, dropdown, date, file, etc.
    required: bool = False
    options: Optional[List[str]] = None  # For dropdown/multi-select
    validation: Optional[dict] = None  # Min, max, regex, etc.
    placeholder: Optional[str] = None
    helpText: Optional[str] = None
    defaultValue: Optional[str] = None
    order: int = 0

class MasterConfiguration(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Machine Master, Process Master, etc.
    description: Optional[str] = None
    icon: Optional[str] = None
    category: str  # Production, Material, Quality, HR, etc.
    fields: List[FieldConfig]
    enableExcelUpload: bool = True
    enableImageUpload: bool = False
    created_at: Optional[datetime] = None
    created_by: Optional[str] = None
    updated_at: Optional[datetime] = None
    updated_by: Optional[str] = None

# API Routes for Master Configuration
@api_router.post("/master-configs")
async def create_master_config(config: MasterConfiguration, current_user: User = Depends(get_current_user)):
    """Create a new master configuration"""
    try:
        config.created_by = current_user.username
        config.created_at = datetime.now(timezone.utc)
        doc = config.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        
        await db.master_configurations.insert_one(doc)
        
        return {
            "message": "Master configuration created successfully",
            "id": config.id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating master config: {str(e)}")

@api_router.get("/master-configs")
async def get_master_configs(category: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get all master configurations"""
    try:
        query = {}
        if category:
            query["category"] = category
        
        configs = await db.master_configurations.find(query, {"_id": 0}).to_list(1000)
        
        for config in configs:
            if isinstance(config.get('created_at'), str):
                config['created_at'] = datetime.fromisoformat(config['created_at'])
            if isinstance(config.get('updated_at'), str):
                config['updated_at'] = datetime.fromisoformat(config['updated_at'])
        
        return configs
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching master configs: {str(e)}")

@api_router.get("/master-configs/{config_id}")
async def get_master_config(config_id: str, current_user: User = Depends(get_current_user)):
    """Get specific master configuration"""
    try:
        config = await db.master_configurations.find_one({"id": config_id}, {"_id": 0})
        
        if not config:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        if isinstance(config.get('created_at'), str):
            config['created_at'] = datetime.fromisoformat(config['created_at'])
        if isinstance(config.get('updated_at'), str):
            config['updated_at'] = datetime.fromisoformat(config['updated_at'])
        
        return config
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching master config: {str(e)}")

@api_router.put("/master-configs/{config_id}")
async def update_master_config(config_id: str, config: MasterConfiguration, current_user: User = Depends(get_current_user)):
    """Update master configuration"""
    try:
        existing = await db.master_configurations.find_one({"id": config_id})
        
        if not existing:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        config.updated_at = datetime.now(timezone.utc)
        config.updated_by = current_user.username
        
        doc = config.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        doc['updated_at'] = doc['updated_at'].isoformat()
        
        await db.master_configurations.update_one(
            {"id": config_id},
            {"$set": doc}
        )
        
        return {"message": "Master configuration updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating master config: {str(e)}")

@api_router.delete("/master-configs/{config_id}")
async def delete_master_config(config_id: str, current_user: User = Depends(get_current_user)):
    """Delete master configuration"""
    try:
        result = await db.master_configurations.delete_one({"id": config_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        # Also delete all data for this master type
        collection_name = f"dynamic_{config_id}"
        await db[collection_name].drop()
        
        return {"message": "Master configuration deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting master config: {str(e)}")

# API Routes for Dynamic Master Data
@api_router.post("/dynamic-masters/{config_id}/data")
async def create_dynamic_master_data(config_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Create data for a dynamic master"""
    try:
        # Get master configuration
        config = await db.master_configurations.find_one({"id": config_id})
        
        if not config:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        # Validate data against configuration
        # (In production, add comprehensive validation based on field configs)
        
        # Add metadata
        data["id"] = str(uuid.uuid4())
        data["created_at"] = datetime.now(timezone.utc).isoformat()
        data["created_by"] = current_user.username
        
        # Store in dynamic collection
        collection_name = f"dynamic_{config_id}"
        await db[collection_name].insert_one(data)
        
        return {
            "message": "Master data created successfully",
            "id": data["id"]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating master data: {str(e)}")

@api_router.get("/dynamic-masters/{config_id}/data")
async def get_dynamic_master_data(config_id: str, current_user: User = Depends(get_current_user)):
    """Get all data for a dynamic master"""
    try:
        # Get master configuration
        config = await db.master_configurations.find_one({"id": config_id})
        
        if not config:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        # Fetch data from dynamic collection
        collection_name = f"dynamic_{config_id}"
        data = await db[collection_name].find({}, {"_id": 0}).to_list(1000)
        
        return data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching master data: {str(e)}")

@api_router.get("/dynamic-masters/{config_id}/data/{data_id}")
async def get_dynamic_master_data_by_id(config_id: str, data_id: str, current_user: User = Depends(get_current_user)):
    """Get specific data item for a dynamic master"""
    try:
        collection_name = f"dynamic_{config_id}"
        data = await db[collection_name].find_one({"id": data_id}, {"_id": 0})
        
        if not data:
            raise HTTPException(status_code=404, detail="Data not found")
        
        return data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching data: {str(e)}")

@api_router.put("/dynamic-masters/{config_id}/data/{data_id}")
async def update_dynamic_master_data(config_id: str, data_id: str, data: dict, current_user: User = Depends(get_current_user)):
    """Update data for a dynamic master"""
    try:
        # Add metadata
        data["updated_at"] = datetime.now(timezone.utc).isoformat()
        data["updated_by"] = current_user.username
        
        # Update in dynamic collection
        collection_name = f"dynamic_{config_id}"
        result = await db[collection_name].update_one(
            {"id": data_id},
            {"$set": data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Data not found")
        
        return {"message": "Master data updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating master data: {str(e)}")

@api_router.delete("/dynamic-masters/{config_id}/data/{data_id}")
async def delete_dynamic_master_data(config_id: str, data_id: str, current_user: User = Depends(get_current_user)):
    """Delete data for a dynamic master"""
    try:
        collection_name = f"dynamic_{config_id}"
        result = await db[collection_name].delete_one({"id": data_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Data not found")
        
        return {"message": "Master data deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting master data: {str(e)}")

# ============================================================================
# MASTER BUILDER INITIALIZATION - Pre-configured Masters
# ============================================================================

PREDEFINED_MASTERS = [
    {
        "id": "buyer_master",
        "name": "Buyer/Customer Master",
        "description": "Manage all buyers and customers",
        "icon": "👥",
        "category": "other",
        "enableExcelUpload": True,
        "enableImageUpload": False,
        "fields": [
            {"id": "1", "name": "name", "label": "Buyer Name", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "contact_person", "label": "Contact Person", "type": "text", "required": False, "order": 1},
            {"id": "3", "name": "email", "label": "Email", "type": "text", "required": False, "order": 2},
            {"id": "4", "name": "phone", "label": "Phone", "type": "text", "required": False, "order": 3},
            {"id": "5", "name": "address", "label": "Address", "type": "textarea", "required": False, "order": 4},
            {"id": "6", "name": "country", "label": "Country", "type": "text", "required": False, "order": 5}
        ]
    },
    {
        "id": "supplier_master",
        "name": "Supplier/Vendor Master",
        "description": "Manage all suppliers and vendors",
        "icon": "🏢",
        "category": "other",
        "enableExcelUpload": True,
        "enableImageUpload": False,
        "fields": [
            {"id": "1", "name": "name", "label": "Supplier Name", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "contact_person", "label": "Contact Person", "type": "text", "required": False, "order": 1},
            {"id": "3", "name": "email", "label": "Email", "type": "text", "required": False, "order": 2},
            {"id": "4", "name": "phone", "label": "Phone", "type": "text", "required": False, "order": 3},
            {"id": "5", "name": "address", "label": "Address", "type": "textarea", "required": False, "order": 4},
            {"id": "6", "name": "material_type", "label": "Material Type", "type": "text", "required": False, "order": 5}
        ]
    },
    {
        "id": "fabric_master",
        "name": "Fabric Master",
        "description": "Manage fabric materials with all specifications",
        "icon": "🧵",
        "category": "material",
        "enableExcelUpload": True,
        "enableImageUpload": True,
        "fields": [
            {"id": "1", "name": "sr_no", "label": "Serial No", "type": "number", "required": False, "order": 0},
            {"id": "2", "name": "fabric_name", "label": "Fabric Name", "type": "text", "required": True, "order": 1},
            {"id": "3", "name": "fabric_type", "label": "Fabric Type", "type": "text", "required": False, "order": 2},
            {"id": "4", "name": "composition", "label": "Composition", "type": "text", "required": False, "order": 3},
            {"id": "5", "name": "gsm", "label": "GSM", "type": "text", "required": False, "order": 4},
            {"id": "6", "name": "width", "label": "Width", "type": "text", "required": False, "order": 5},
            {"id": "7", "name": "color", "label": "Color", "type": "text", "required": False, "order": 6},
            {"id": "8", "name": "supplier", "label": "Supplier", "type": "text", "required": False, "order": 7},
            {"id": "9", "name": "cost_per_unit", "label": "Cost per Unit", "type": "decimal", "required": False, "order": 8},
            {"id": "10", "name": "unit", "label": "Unit", "type": "dropdown", "required": False, "options": ["meter", "kg", "yard"], "order": 9},
            {"id": "11", "name": "final_item", "label": "Final Item", "type": "text", "required": False, "order": 10}
        ]
    },
    {
        "id": "color_master",
        "name": "Color Master",
        "description": "Manage color codes and specifications",
        "icon": "🎨",
        "category": "other",
        "enableExcelUpload": True,
        "enableImageUpload": False,
        "fields": [
            {"id": "1", "name": "name", "label": "Color Name", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "code", "label": "Color Code", "type": "text", "required": False, "order": 1},
            {"id": "3", "name": "hex_value", "label": "Hex Value", "type": "text", "required": False, "order": 2},
            {"id": "4", "name": "pantone", "label": "Pantone Code", "type": "text", "required": False, "order": 3}
        ]
    },
    {
        "id": "size_master",
        "name": "Size Master",
        "description": "Manage size specifications",
        "icon": "📏",
        "category": "other",
        "enableExcelUpload": True,
        "enableImageUpload": False,
        "fields": [
            {"id": "1", "name": "name", "label": "Size Name", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "category", "label": "Category", "type": "dropdown", "required": False, "options": ["Clothing", "Shoes", "Accessories"], "order": 1},
            {"id": "3", "name": "measurements", "label": "Measurements", "type": "text", "required": False, "order": 2}
        ]
    },
    {
        "id": "article_master",
        "name": "Article/Style Master",
        "description": "Manage article and style specifications",
        "icon": "👕",
        "category": "production",
        "enableExcelUpload": True,
        "enableImageUpload": True,
        "fields": [
            {"id": "1", "name": "code", "label": "Article Code", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "name", "label": "Article Name", "type": "text", "required": True, "order": 1},
            {"id": "3", "name": "description", "label": "Description", "type": "textarea", "required": False, "order": 2},
            {"id": "4", "name": "category", "label": "Category", "type": "text", "required": False, "order": 3},
            {"id": "5", "name": "season", "label": "Season", "type": "dropdown", "required": False, "options": ["Spring", "Summer", "Fall", "Winter"], "order": 4}
        ]
    },
    {
        "id": "raw_material_master",
        "name": "Raw Material Master",
        "description": "Manage raw materials and components",
        "icon": "📦",
        "category": "material",
        "enableExcelUpload": True,
        "enableImageUpload": False,
        "fields": [
            {"id": "1", "name": "name", "label": "Material Name", "type": "text", "required": True, "order": 0},
            {"id": "2", "name": "type", "label": "Material Type", "type": "dropdown", "required": False, "options": ["Fabric", "Trim", "Accessory", "Chemical"], "order": 1},
            {"id": "3", "name": "supplier", "label": "Supplier", "type": "text", "required": False, "order": 2},
            {"id": "4", "name": "unit_price", "label": "Unit Price", "type": "decimal", "required": False, "order": 3},
            {"id": "5", "name": "unit", "label": "Unit", "type": "text", "required": False, "order": 4}
        ]
    }
]

@api_router.post("/initialize-predefined-masters")
async def initialize_predefined_masters(current_user: User = Depends(get_current_user)):
    """Initialize pre-configured masters for existing data types"""
    try:
        created_count = 0
        skipped_count = 0
        migrated_count = 0
        
        for master_config in PREDEFINED_MASTERS:
            # Check if master already exists
            existing = await db.master_configurations.find_one({"id": master_config["id"]})
            
            if not existing:
                # Create the master configuration
                config_doc = {
                    **master_config,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": current_user.username
                }
                await db.master_configurations.insert_one(config_doc)
                created_count += 1
                
                # Migrate existing data if any
                old_collection_map = {
                    "buyer_master": "buyers",
                    "supplier_master": "suppliers",
                    "fabric_master": "fabrics",
                    "color_master": "colors",
                    "size_master": "sizes",
                    "article_master": "articles",
                    "raw_material_master": "raw_materials"
                }
                
                old_collection = old_collection_map.get(master_config["id"])
                if old_collection:
                    # Get data from old collection
                    old_data = await db[old_collection].find({}, {"_id": 0}).to_list(1000)
                    
                    if old_data:
                        # Migrate to dynamic collection
                        new_collection = f"dynamic_{master_config['id']}"
                        for item in old_data:
                            item["created_at"] = datetime.now(timezone.utc).isoformat()
                            item["created_by"] = "system_migration"
                            if "id" not in item:
                                item["id"] = str(uuid.uuid4())
                        
                        await db[new_collection].insert_many(old_data)
                        migrated_count += len(old_data)
            else:
                skipped_count += 1
        
        return {
            "message": "Pre-configured masters initialized",
            "created": created_count,
            "skipped": skipped_count,
            "data_migrated": migrated_count
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error initializing masters: {str(e)}")

@api_router.get("/predefined-masters/status")
async def get_predefined_masters_status(current_user: User = Depends(get_current_user)):
    """Check which predefined masters are already initialized"""
    try:
        status = []
        for master in PREDEFINED_MASTERS:
            existing = await db.master_configurations.find_one({"id": master["id"]})
            
            # Count data in both old and new collections
            old_collection_map = {
                "buyer_master": "buyers",
                "supplier_master": "suppliers", 
                "fabric_master": "fabrics",
                "color_master": "colors",
                "size_master": "sizes",
                "article_master": "articles",
                "raw_material_master": "raw_materials"
            }
            
            old_collection = old_collection_map.get(master["id"])
            old_count = 0
            new_count = 0
            
            if old_collection:
                old_count = await db[old_collection].count_documents({})
            
            if existing:
                new_collection = f"dynamic_{master['id']}"
                new_count = await db[new_collection].count_documents({})
            
            status.append({
                "id": master["id"],
                "name": master["name"],
                "initialized": existing is not None,
                "old_data_count": old_count,
                "new_data_count": new_count
            })
        
        return status
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error checking status: {str(e)}")

        raise HTTPException(status_code=500, detail=f"Error deleting master data: {str(e)}")

# Bulk Excel Upload for Dynamic Masters
@api_router.post("/dynamic-masters/{config_id}/bulk-upload")
async def bulk_upload_dynamic_master(
    config_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Bulk upload data for dynamic master via Excel"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file")
        
        # Get master configuration
        config = await db.master_configurations.find_one({"id": config_id})
        
        if not config:
            raise HTTPException(status_code=404, detail="Master configuration not found")
        
        # Read Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        
        # Process rows
        added_count = 0
        errors = []
        collection_name = f"dynamic_{config_id}"
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        field_name = headers[idx]
                        data[field_name] = value if value is not None else ""
                
                # Add metadata
                data["id"] = str(uuid.uuid4())
                data["created_at"] = datetime.now(timezone.utc).isoformat()
                data["created_by"] = current_user.username
                
                await db[collection_name].insert_one(data)
                added_count += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        return {
            "message": f"Successfully uploaded {added_count} records",
            "added_count": added_count,
            "errors": errors
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading data: {str(e)}")

# Excel Upload Route
@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        
        results = {
            "colors_added": 0,
            "articles_added": 0,
            "sizes_added": 0,
            "raw_materials_added": 0,
            "fabrics_added": 0,
            "errors": []
        }
        
        # Process Color ID sheet
        if "Color ID" in workbook.sheetnames:
            sheet = workbook["Color ID"]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[0] and row[1]:
                        color_data = row[1].split("/")
                        if len(color_data) >= 2:
                            color_id = color_data[0].strip()
                            color_name = color_data[1].strip()
                            
                            # Check if color already exists
                            existing = await db.colors.find_one({"code": color_id})
                            if not existing:
                                color_obj = Color(
                                    name=color_name,
                                    code=color_id,
                                    hex_value=None
                                )
                                doc = color_obj.model_dump()
                                doc['created_at'] = doc['created_at'].isoformat()
                                await db.colors.insert_one(doc)
                                results["colors_added"] += 1
                except Exception as e:
                    results["errors"].append(f"Color sheet row {row_idx}: {str(e)}")
        
        # Process Art No. sheet
        if "Art No." in workbook.sheetnames:
            sheet = workbook["Art No."]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[0]:
                        article_code = str(row[0]).strip()
                        
                        # Check if article already exists
                        existing = await db.articles.find_one({"code": article_code})
                        if not existing:
                            article_obj = Article(
                                name=article_code,
                                code=article_code,
                                description=f"Article {article_code}",
                                buyer_id=None
                            )
                            doc = article_obj.model_dump()
                            doc['created_at'] = doc['created_at'].isoformat()
                            await db.articles.insert_one(doc)
                            results["articles_added"] += 1
                except Exception as e:
                    results["errors"].append(f"Article sheet row {row_idx}: {str(e)}")
        
        # Process Units Master sheet
        if "Units Master" in workbook.sheetnames:
            sheet = workbook["Units Master"]
            sort_order = 1
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[0]:
                        unit_name = str(row[0]).strip()
                        
                        # Check if size already exists
                        existing = await db.sizes.find_one({"code": unit_name})
                        if not existing:
                            size_obj = Size(
                                name=unit_name,
                                code=unit_name,
                                sort_order=sort_order
                            )
                            doc = size_obj.model_dump()
                            doc['created_at'] = doc['created_at'].isoformat()
                            await db.sizes.insert_one(doc)
                            results["sizes_added"] += 1
                            sort_order += 1
                except Exception as e:
                    results["errors"].append(f"Units sheet row {row_idx}: {str(e)}")
        
        # Process Components sheet (as Raw Materials)
        if "Components" in workbook.sheetnames:
            sheet = workbook["Components"]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if row[0]:
                        component_name = str(row[0]).strip()
                        component_code = component_name[:20].upper().replace(" ", "_")
                        
                        # Check if material already exists
                        existing = await db.raw_materials.find_one({"code": component_code})
                        if not existing:
                            material_obj = RawMaterial(
                                name=component_name,
                                code=component_code,
                                material_type="accessories",
                                unit="pieces",
                                cost_per_unit=0.0,
                                supplier_id=None
                            )
                            doc = material_obj.model_dump()
                            doc['created_at'] = doc['created_at'].isoformat()
                            await db.raw_materials.insert_one(doc)
                            results["raw_materials_added"] += 1
                except Exception as e:
                    results["errors"].append(f"Components sheet row {row_idx}: {str(e)}")
        
        # Process Fabric Master Data sheet
        if "FABRIC MASTER DATA" in workbook.sheetnames:
            sheet = workbook["FABRIC MASTER DATA"]
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Skip empty rows
                    if not row[0] or str(row[0]).strip() == "":
                        continue
                    
                    # Helper function to clean values
                    def clean_value(val):
                        if val is None:
                            return ""
                        val_str = str(val).strip()
                        # Handle Excel "None" strings
                        if val_str.upper() == "NONE" or val_str == "":
                            return ""
                        return val_str
                    
                    # Parse values with proper null handling
                    item_type = clean_value(row[0])
                    count_const = clean_value(row[1])
                    fabric_name = clean_value(row[2])
                    composition = clean_value(row[3])
                    add_description = clean_value(row[4])
                    
                    # Handle GSM - can be numeric or None
                    gsm = None
                    if row[5] and str(row[5]).strip().upper() != "NONE":
                        try:
                            gsm = int(float(row[5]))
                        except (ValueError, TypeError):
                            pass
                    
                    width = clean_value(row[6])
                    color = clean_value(row[7])
                    final_item = clean_value(row[8])
                    avg_roll_size = clean_value(row[9])
                    unit = clean_value(row[10]) or "Pcs"
                    
                    fabric_obj = Fabric(
                        item_type=item_type,
                        count_const=count_const,
                        fabric_name=fabric_name,
                        composition=composition,
                        add_description=add_description,
                        gsm=gsm,
                        width=width if width else None,
                        color=color if color else None,
                        final_item=final_item,
                        avg_roll_size=avg_roll_size if avg_roll_size else None,
                        unit=unit
                    )
                    doc = fabric_obj.model_dump()
                    doc['created_at'] = doc['created_at'].isoformat()
                    await db.fabrics.insert_one(doc)
                    results["fabrics_added"] += 1
                except Exception as e:
                    results["errors"].append(f"Fabric sheet row {row_idx}: {str(e)}")
        
        return {
            "message": "Excel file processed successfully",
            "results": results
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

# Image Upload Route
@api_router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    try:
        # Create uploads directory if it doesn't exist
        upload_dir = Path("/app/uploads")
        upload_dir.mkdir(exist_ok=True)
        
        # Generate unique filename
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        unique_filename = f"{uuid.uuid4()}.{file_extension}"
        file_path = upload_dir / unique_filename
        
        # Save file
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        
        # Return the URL path
        image_url = f"/uploads/{unique_filename}"
        
        return {
            "message": "Image uploaded successfully",
            "image_url": image_url,
            "filename": unique_filename
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading image: {str(e)}")

# Include router

# BOM Form Configuration Routes
@api_router.get("/bom-form-config")
async def get_bom_form_config():
    """Get BOM form configuration"""
    config = await db.bom_form_configs.find_one({"type": "dyeing_bom"})
    if not config:
        return {}
    config.pop("_id", None)
    return config

@api_router.post("/bom-form-config")
async def save_bom_form_config(config: dict):
    """Save BOM form configuration"""
    try:
        config["type"] = "dyeing_bom"
        config["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.bom_form_configs.update_one(
            {"type": "dyeing_bom"},
            {"$set": config},
            upsert=True
        )
        
        return {"message": "Configuration saved successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving configuration: {str(e)}")

app.include_router(api_router)

# Mount static files for image serving
app.mount("/uploads", StaticFiles(directory="/app/uploads"), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
